"""
Plugin de Exemplo: Exportador para Excel

Este plugin demonstra como criar funcionalidades customizadas
para o SimpleNFE. Ele adiciona a capacidade de exportar itens
extraídos para formato Excel (.xlsx).

Autor: Comunidade SimpleNFE
Versão: 1.0.0
"""

from plugins import BasePlugin
from typing import Dict, Any
import json

class ExportadorExcelPlugin(BasePlugin):
    """Plugin que exporta itens extraídos para Excel"""
    
    @property
    def name(self) -> str:
        return "Exportador Excel"
    
    @property
    def version(self) -> str:
        return "1.0.0"
    
    @property
    def description(self) -> str:
        return "Exporta itens extraídos para planilha Excel (.xlsx) com formatação profissional"
    
    @property
    def author(self) -> str:
        return "Comunidade SimpleNFE"
    
    def initialize(self, app_context: Dict[str, Any]) -> bool:
        """Inicializa o plugin"""
        self.app = app_context.get('app')
        self.items = app_context.get('extracted_items', [])
        
        # Verifica se openpyxl está instalado
        try:
            import openpyxl
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            print("⚠️ Plugin Exportador Excel: openpyxl não instalado. Execute: pip install openpyxl")
        
        return True
    
    def get_menu_label(self) -> str:
        return "📊 Exportar para Excel"
    
    def get_toolbar_icon(self) -> str:
        return "📊"
    
    def execute(self, **kwargs) -> Any:
        """Executa a exportação para Excel"""
        if not self.openpyxl_available:
            return {
                'success': False,
                'message': 'openpyxl não está instalado. Execute: pip install openpyxl'
            }
        
        import tkinter as tk
        from tkinter import filedialog, messagebox
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Atualiza itens do contexto
        self.items = kwargs.get('items', self.items)
        
        if not self.items:
            return {
                'success': False,
                'message': 'Nenhum item disponível para exportar'
            }
        
        # Diálogo para salvar arquivo
        filepath = filedialog.asksaveasfilename(
            title="Exportar para Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        
        if not filepath:
            return {'success': False, 'message': 'Exportação cancelada'}
        
        try:
            # Cria workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Itens NF-e"
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalhos
            headers = ["Descrição", "Quantidade", "Unidade", "Valor Unit.", "Valor Total", "Fornecedor", "Data NF-e"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Dados
            for row, item in enumerate(self.items, 2):
                ws.cell(row, 1, item.get('descricao', '')).border = border
                ws.cell(row, 2, item.get('quantidade', 0)).border = border
                ws.cell(row, 3, item.get('unidade', '')).border = border
                ws.cell(row, 4, item.get('valor_unitario', 0)).border = border
                ws.cell(row, 5, item.get('valor_total', 0)).border = border
                ws.cell(row, 6, item.get('fornecedor', '')).border = border
                ws.cell(row, 7, item.get('data_nota', '')).border = border
                
                # Formata valores monetários
                ws.cell(row, 4).number_format = 'R$ #,##0.00'
                ws.cell(row, 5).number_format = 'R$ #,##0.00'
            
            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 15
            
            # Congela primeira linha
            ws.freeze_panes = 'A2'
            
            # Salva
            wb.save(filepath)
            
            messagebox.showinfo(
                "Sucesso",
                f"✅ Planilha Excel exportada com sucesso!\n\n{filepath}\n\n{len(self.items)} itens exportados"
            )
            
            return {
                'success': True,
                'message': f'Exportado {len(self.items)} itens para {filepath}',
                'filepath': filepath
            }
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel:\n{str(e)}")
            return {'success': False, 'message': str(e)}
    
    def get_settings_ui(self) -> Dict[str, Any]:
        """Define configurações do plugin"""
        return {
            'campos': [
                {
                    'nome': 'incluir_totais',
                    'tipo': 'checkbox',
                    'label': 'Incluir linha de totais',
                    'default': True
                },
                {
                    'nome': 'congelar_painel',
                    'tipo': 'checkbox',
                    'label': 'Congelar primeira linha',
                    'default': True
                }
            ]
        }
